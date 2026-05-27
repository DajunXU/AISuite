import os
import json
import asyncio
from datetime import datetime
from typing import List, Dict, Any
from pathlib import Path
from sqlalchemy.orm import Session
from sqlalchemy import text
from fastapi import UploadFile, HTTPException, status
import aiofiles

from ..models.knowledge import UploadedFile, DocumentChunk, KnowledgeBase
from ..models.llm import LLMModel, LLMKeyManager
from ..models.embedding import EmbeddingModel, EmbeddingKeyManager
from ..core.config import settings
from ..core.logger import get_logger
from ..services.rag import RAGService

logger = get_logger("file_processor")


class FileProcessor:
    """文件处理器 - 负责文件上传、解析和向量化"""
    
    def __init__(self):
        self.upload_dir = Path(settings.UPLOAD_DIR)
        self.upload_dir.mkdir(parents=True, exist_ok=True)
    
    async def save_uploaded_file(self, file: UploadFile, knowledge_base_id: int) -> UploadedFile:
        """保存上传的文件"""
        # 生成唯一文件名
        file_extension = Path(file.filename).suffix
        unique_filename = f"{knowledge_base_id}_{int(datetime.now().timestamp())}{file_extension}"
        file_path = self.upload_dir / unique_filename
        
        # 保存文件
        async with aiofiles.open(file_path, 'wb') as f:
            content = await file.read()
            await f.write(content)
        
        # 获取文件大小
        file_size = len(content)
        
        return UploadedFile(
            knowledge_base_id=knowledge_base_id,
            filename=unique_filename,
            original_filename=file.filename,
            file_path=str(file_path),
            file_size=file_size,
            file_type=file.content_type or "application/octet-stream",
            status="uploaded"
        )
    
    async def process_file(self, uploaded_file: UploadedFile, db: Session) -> bool:
        """处理上传的文件（解析和向量化）"""
        try:
            # 更新文件状态为处理中
            uploaded_file.status = "processing"
            db.commit()
            
            # 根据文件类型选择解析器
            file_path = Path(uploaded_file.file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"文件不存在: {file_path}")
            
            # 解析文件内容
            content = await self._parse_file(file_path, uploaded_file.file_type)
            
            # 分块处理
            chunks = self._chunk_content(content)
            
            # 向量化处理
            await self._vectorize_chunks(chunks, uploaded_file.knowledge_base_id, db)
            
            # 更新文件状态为已处理
            uploaded_file.status = "processed"
            uploaded_file.vectorized_at = datetime.now()
            db.commit()
            
            return True
            
        except Exception as e:
            # 更新文件状态为错误
            uploaded_file.status = "error"
            db.commit()
            logger.error(f"文件处理错误: {e}")
            return False
    
    async def _parse_file(self, file_path: Path, file_type: str) -> str:
        """解析文件内容"""
        # 根据文件类型选择解析方式
        if file_type == "text/plain" or file_path.suffix == ".txt":
            async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
                return await f.read()
        
        elif file_type == "application/pdf" or file_path.suffix == ".pdf":
            try:
                import pdfplumber
                text = ""
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
                return text
            except ImportError:
                try:
                    import PyPDF2
                    with open(file_path, 'rb') as f:
                        pdf_reader = PyPDF2.PdfReader(f)
                        text = ""
                        for page in pdf_reader.pages:
                            text += page.extract_text() or ""
                        return text
                except ImportError:
                    raise Exception("PDF解析需要安装pdfplumber或PyPDF2: pip install pdfplumber PyPDF2")
        
        elif file_type in ["application/vnd.openxmlformats-officedocument.wordprocessingml.document", 
                          "application/msword"] or file_path.suffix in [".docx", ".doc"]:
            if file_path.suffix == ".doc":
                try:
                    import textract
                    text = textract.process(str(file_path)).decode('utf-8')
                    if text.strip():
                        return text
                except ImportError:
                    if settings.SYSTEM_PLATFORM == "Windows":
                        raise Exception(".doc 格式解析失败，请将文件转换为 .docx 格式后重新上传")
                    else:
                        raise Exception(".doc 格式解析失败，请安装antiword: sudo apt-get install antiword (Linux)")
                except Exception as e:
                    if "antiword" in str(e) or "exit code" in str(e):
                        if settings.SYSTEM_PLATFORM == "Windows":
                            raise Exception(".doc 格式解析失败，请将文件转换为 .docx 格式后重新上传")
                        else:
                            raise Exception(".doc 格式解析失败，请安装antiword: sudo apt-get install antiword")
                    raise Exception(f".doc 格式解析失败: {str(e)}")
            
            try:
                import docx
                doc = docx.Document(file_path)
                text = ""
                for paragraph in doc.paragraphs:
                    text += paragraph.text + "\n"
                if text.strip():
                    return text
            except ImportError:
                raise Exception("Word文档解析需要安装python-docx: pip install python-docx")
            except Exception as docx_error:
                raise Exception(f"Word文档解析失败: {str(docx_error)}")
            
            raise Exception("Word文档解析后内容为空")
        
        elif file_type in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                          "application/vnd.ms-excel"] or file_path.suffix in [".xlsx", ".xls"]:
            try:
                import openpyxl
                text = ""
                wb = openpyxl.load_workbook(file_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    text += f"\n--- Sheet: {sheet_name} ---\n"
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                        if row_text.strip():
                            text += row_text + "\n"
                wb.close()
                if text.strip():
                    return text
            except ImportError:
                raise Exception("Excel文件解析需要安装openpyxl: pip install openpyxl")
            except Exception as e:
                raise Exception(f"Excel文件解析失败: {str(e)}")
            
            raise Exception("Excel文件解析后内容为空")
        
        elif file_type == "text/markdown" or file_path.suffix == ".md":
            async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
                return await f.read()
        
        elif file_type == "text/csv" or file_path.suffix == ".csv":
            import csv
            text = ""
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                for row in reader:
                    text += " | ".join(row) + "\n"
            if text.strip():
                return text
            raise Exception("CSV文件解析后内容为空")
        
        else:
            # 默认按文本文件处理
            async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
                return await f.read()
    
    def _chunk_content(self, content: str, chunk_size: int = 1000) -> List[str]:
        """将内容分块"""
        chunks = []
        for i in range(0, len(content), chunk_size):
            chunk = content[i:i + chunk_size]
            chunks.append(chunk)
        return chunks
    
    async def _vectorize_chunks(self, chunks: List[str], knowledge_base_id: int, db: Session):
        """向量化文本块"""
        
        rag_service = RAGService(db)
        
        # 获取知识库配置的向量模型
        knowledge_base = db.query(KnowledgeBase).filter(KnowledgeBase.id == knowledge_base_id).first()
        
        if knowledge_base and knowledge_base.embedding_model_id:
            # 使用知识库配置的向量模型
            embedding_model = db.query(EmbeddingModel).filter(EmbeddingModel.id == knowledge_base.embedding_model_id).first()
            if embedding_model:
                try:
                    key_manager = EmbeddingKeyManager()
                    api_key = ""
                    if embedding_model.api_key_encrypted:
                        api_key = key_manager.decrypt_api_key(embedding_model.api_key_encrypted)
                    if not api_key:
                        if embedding_model.provider == "qwen" or "dashscope" in (embedding_model.base_url or ""):
                            import os
                            api_key = os.environ.get("DASHSCOPE_API_KEY", "")
                        elif embedding_model.provider == "zhipu":
                            import os
                            api_key = os.environ.get("ZHIPU_API_KEY", "")
                    
                    rag_service._init_embeddings_from_model(embedding_model)
                    logger.info(f"文件向量化使用知识库配置的向量模型: {embedding_model.name}, 提供商: {embedding_model.provider}")
                except Exception as e:
                    logger.error(f"初始化知识库向量模型失败: {e}，尝试使用默认模型")
                    rag_service._init_embeddings()
            else:
                logger.warning(f"未找到向量模型 ID {knowledge_base.embedding_model_id}，使用默认配置")
                rag_service._init_embeddings()
        else:
            # 获取默认 LLM 模型配置（兼容旧逻辑）
            default_model = db.query(LLMModel).filter(
                LLMModel.is_default == True,
                LLMModel.is_active == True
            ).first()
            
            if default_model:
                try:
                    key_manager = LLMKeyManager()
                    api_key = key_manager.decrypt_api_key(default_model.api_key_encrypted)
                    rag_service._init_llm(
                        api_key=api_key,
                        base_url=default_model.base_url,
                        provider=default_model.provider,
                        model=default_model
                    )
                    logger.info(f"文件向量化使用默认LLM模型: {default_model.name}, 提供商: {default_model.provider}")
                except Exception as e:
                    logger.error(f"初始化AI客户端失败: {e}")
                    rag_service._init_llm()
            else:
                logger.warning("没有找到默认模型配置，使用环境默认密钥")
                rag_service._init_llm()
        
        # 获取向量维度 - 从 embedding_model 动态获取，如果没有则查询默认模型
        vector_dimensions = None
        if knowledge_base and knowledge_base.embedding_model_id:
            embedding_model = db.query(EmbeddingModel).filter(EmbeddingModel.id == knowledge_base.embedding_model_id).first()
            if embedding_model and embedding_model.dimensions:
                vector_dimensions = embedding_model.dimensions
                logger.info(f"使用知识库配置的向量维度: {vector_dimensions}")
        
        # 如果没有获取到维度，查询默认向量模型
        if vector_dimensions is None:
            default_embedding_model = db.query(EmbeddingModel).filter(
                EmbeddingModel.is_default == True,
                EmbeddingModel.is_active == True
            ).first()
            if default_embedding_model and default_embedding_model.dimensions:
                vector_dimensions = default_embedding_model.dimensions
                logger.info(f"使用默认向量维度: {vector_dimensions}")
        
        # 如果仍然没有维度，抛出错误而不是使用错误的默认值
        if vector_dimensions is None:
            raise ValueError("无法确定向量维度，请检查向量模型配置")
        
        for i, chunk in enumerate(chunks):
            # 生成向量
            try:
                embedding = await rag_service.generate_embedding(chunk)
            except Exception as e:
                logger.error(f"生成嵌入向量失败: {e}")
                embedding = [0.0] * vector_dimensions
            
            # 创建文档块记录 - 使用CAST将列表转换为vector
            try:
                # 使用PostgreSQL CAST语法将数组转换为vector
                
                embedding_str = "[" + ",".join(map(str, embedding)) + "]"
                
                # 直接使用原生SQL插入
                db.execute(
                    text(f"""
                        INSERT INTO document_chunks 
                        (knowledge_base_id, content, embedding, chunk_metadata, chunk_index, created_at)
                        VALUES (:kb_id, :content, CAST(:embedding AS vector({vector_dimensions})), :metadata, :idx, NOW())
                    """),
                    {
                        "kb_id": knowledge_base_id,
                        "content": chunk,
                        "embedding": embedding_str,
                        "metadata": json.dumps({"chunk_index": i, "chunk_size": len(chunk)}),
                        "idx": i
                    }
                )
            except Exception as e:
                logger.error(f"创建文档块失败: {e}")
                continue
        
        db.commit()


# 全局文件处理器实例
file_processor = FileProcessor()